import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r"D:\study\pythonstudy.xlsx")

# Getting sheets from the workbook

print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

mySheet = wb.create_sheet("mySheet")
print(wb.sheetnames)

# sheet3 = wb.get_sheet_by_name("sheet3")
# sheet4 = wb["mySheet"]
ws = wb.active
print(ws)
print(ws["A1"])
print(ws["A1"].value)

# c = ws["B1"]
# print("Row {},column {} is {}".format(c.row,c.column,c.value))
# print("Cell {} is{}".format(c.coordinate,c.value))
#
# print(ws.cell(row=1,column=2))
# print(ws.cell(row=1,column=2).value)
# for i in range(1,8,2):
#     print(i,ws.cell(row=1,column=2).value)
# colC = ws["C"]
# row6 = ws[6]
# col_range = ws["B:C"]
# row_range = ws[2:6]
# for col in col_range:
#     for cell in col:
#         print(cell.value)
# for row in row_range:
#     for cell in row:
#         print(cell.value)

# for row in ws.iter_rows(min_row=1,max_row=2,max_col=2):
#     for cell in row:
#         print(cell)
# cell_range = ws["A1:C3"]
# for rowOfCellObjects in cell_range:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate,cellObj.value)
#     print("————————End of Row ————————")
#print("{} * {}".format(ws.max_row,ws.max_column))
print(get_column_letter(2),get_column_letter(45),get_column_letter(900))
